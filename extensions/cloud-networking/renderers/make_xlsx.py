#!/usr/bin/env python3
"""make_xlsx.py - Cloud Networking Excel workbook renderer (with REAL formulas).

Install: pip install openpyxl
Usage:   python make_xlsx.py --spec spec.json --output report.xlsx

Spec format (JSON) - multi-sheet workbook with real formulas:
{
  "title": "IP Capacity Plan - 2026-05-27",
  "specialist": "Capacity Planner",
  "named_ranges": [
    {"name": "total_prefix_bits", "ref": "Inputs!$B$2"},
    {"name": "subnet_prefix_bits", "ref": "Inputs!$B$3"}
  ],
  "sheets": [
    {
      "name": "Summary",
      "cells": [
        {"ref": "A1", "value": "Headline", "style": "h1"},
        {"ref": "A3", "value": "Usable hosts / subnet"},
        {"ref": "B3", "formula": "=2^(32-subnet_prefix_bits)-5", "number_format": "#,##0"}
      ]
    },
    {
      "name": "Inputs",
      "headers": ["Parameter", "Value"],
      "rows": [["VNet prefix (/n)", 16], ["Subnet prefix (/n)", 24]]
    }
  ]
}

The policy mandates REAL formulas (=SUM, =IF, =NPV, =FV, =PMT, =2^(...)) - never
hard-coded computed values - so the user can edit one Inputs cell and see the
Summary / Scenarios sheets update automatically.

NAMED RANGES (required for any spec with computed Summary cells): declare
inputs at the top of the spec and reference them by name in formulas.
This eliminates the off-by-one A1-coordinate class of bug (a formula pointing
at the wrong Inputs cell silently ships a wrong-by-X% number). The validator
(validate_named_ranges) rejects malformed refs, missing sheets, duplicate names,
and invalid Excel names. Example specs live in renderers/examples/.
"""
from __future__ import annotations
import argparse, json, pathlib, re, sys

PRIMARY = "0E5A9C"
STRIPE = "F3F7FB"


def style_header(cell):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_h1(cell):
    from openpyxl.styles import Font
    cell.font = Font(name="Segoe UI", bold=True, color=PRIMARY, size=14)


def style_body(cell):
    from openpyxl.styles import Font
    cell.font = Font(name="Segoe UI", size=10)


def build_sheet(wb, spec: dict) -> None:
    ws = wb.create_sheet(title=spec["name"][:31])

    if "headers" in spec and "rows" in spec:
        for col_idx, h in enumerate(spec["headers"], start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            style_header(c)
        for row_idx, row in enumerate(spec["rows"], start=2):
            for col_idx, val in enumerate(row, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                style_body(c)
                if isinstance(val, (int, float)) and col_idx > 1:
                    if abs(val) < 1 and val != 0:
                        c.number_format = "0.00%"
                    else:
                        c.number_format = "#,##0"
        ws.freeze_panes = "A2"
        for col_idx in range(1, len(spec["headers"]) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    for cell_spec in spec.get("cells", []):
        ref = cell_spec["ref"]
        c = ws[ref]
        if "formula" in cell_spec:
            c.value = cell_spec["formula"]
        elif "value" in cell_spec:
            c.value = cell_spec["value"]
        nf = cell_spec.get("number_format")
        if nf:
            c.number_format = nf
        style_name = cell_spec.get("style", "body")
        if style_name == "h1":
            style_h1(c)
        elif style_name == "header":
            style_header(c)
        else:
            style_body(c)


# Matches "SheetName!$B$4" or "Sheet Name!B4:C10" with optional $ anchors.
# Quoted sheet names ("'My Sheet'!$B$4") are tolerated by stripping quotes.
_NAMED_RANGE_REF_RE = re.compile(
    r"""^
        '?(?P<sheet>[^'!]+?)'?      # sheet name, optionally single-quoted
        !
        \$?[A-Z]+\$?\d+             # start cell, optional anchors
        (?::\$?[A-Z]+\$?\d+)?       # optional range end
        $""",
    re.VERBOSE,
)


def validate_named_ranges(spec: dict) -> None:
    """Validate that every named_range targets a sheet declared in the spec.

    Catches the off-by-sheet class of bug (e.g., a name pointing at
    'Imputs!$B$4' or at a sheet name dropped between drafts). Raised
    errors are loud so the agent fixes the JSON before shipping a
    silently-broken workbook.
    """
    names = spec.get("named_ranges", [])
    if not names:
        return
    sheet_titles = {s["name"] for s in spec.get("sheets", [])}
    seen: set[str] = set()
    for entry in names:
        if "name" not in entry or "ref" not in entry:
            raise ValueError(
                f"named_ranges entry missing 'name' or 'ref': {entry!r}"
            )
        nm = entry["name"]
        ref = entry["ref"]
        if nm in seen:
            raise ValueError(f"named_ranges contains duplicate name {nm!r}")
        seen.add(nm)
        if not re.match(r"^[A-Za-z_][A-Za-z0-9_.]*$", nm):
            raise ValueError(
                f"named_range {nm!r} is not a valid Excel name "
                f"(must start with a letter or underscore; letters / digits / _ / . only)"
            )
        m = _NAMED_RANGE_REF_RE.match(ref)
        if not m:
            raise ValueError(
                f"named_range {nm!r} has malformed ref {ref!r}; "
                f"expected e.g. 'Inputs!$B$4' or 'Analysis!B2:B40'"
            )
        sheet = m.group("sheet")
        if sheet not in sheet_titles:
            raise ValueError(
                f"named_range {nm!r} references sheet {sheet!r} "
                f"which is not declared in spec.sheets "
                f"(known: {sorted(sheet_titles)})"
            )


def apply_named_ranges(wb, spec: dict) -> None:
    """Emit each spec named_range as an openpyxl DefinedName at workbook scope."""
    from openpyxl.workbook.defined_name import DefinedName

    for entry in spec.get("named_ranges", []):
        dn = DefinedName(name=entry["name"], attr_text=entry["ref"])
        wb.defined_names[entry["name"]] = dn


def main() -> int:
    ap = argparse.ArgumentParser(description="Render a Cloud Networking multi-sheet XLSX workbook with real formulas.")
    ap.add_argument("--spec", required=True, type=pathlib.Path, help="JSON spec describing sheets / cells / formulas")
    ap.add_argument("--output", required=True, type=pathlib.Path)
    args = ap.parse_args()

    try:
        from openpyxl import Workbook
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

    spec = json.loads(args.spec.read_text(encoding="utf-8"))
    validate_named_ranges(spec)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_spec in spec.get("sheets", []):
        build_sheet(wb, sheet_spec)

    apply_named_ranges(wb, spec)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="Summary")
        ws["A1"] = spec.get("title", "Cloud Networking Report")
        style_h1(ws["A1"])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    size = args.output.stat().st_size
    sheet_count = len(wb.sheetnames)
    print(f"OK  {args.output}  ({size:,} bytes, {sheet_count} sheet(s))")
    return 0


if __name__ == "__main__":
    sys.exit(main())
